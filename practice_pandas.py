#pip install pandas
#pip install Pyarrow
import pandas as pd
import pyautogui as pyautogui #pip install pyautogui
import openpyxl

#data = pd.read_excel("C:/study/CS codes.xlsx")
#print(data)

# cs_codes = pd.read_excel("C:/study/CS codes.xlsx", usecols=[1])
# cs_codes = cs_codes.dropna().values.tolist() 
# #.dropna(): remove NaN values
# #.values.tolist(): convert as a list

# #print(cs_codes)

# # print(list[0])
# # print(f"{list[0]}")

# cs_codes = [int(float(str(x).split('.')[0])) for sublist in cs_codes for x in sublist]

# #print(cs_codes)
# #print(type(converted_list[0]))


# # print(type(list[0]))

# coupon_code = pd.read_excel("C:/study/CS codes.xlsx", usecols=[1], sheet_name="coupon")
# print(coupon_code.shape)
# #print(coupon_code.loc[0])
# #print(type(coupon_code.tail(1)))
# #print(coupon_code.tail(1).to_string, "aaaaa")
# print(type(coupon_code.iat[0,0]))
# #print(coupon_code.values.tolist()[1])
# #https://velog.io/@cha-suyeon/%ED%8C%90%EB%8B%A4%EC%8A%A4pandas-%EC%8B%A4%EC%8A%B5-%EB%8D%B0%EC%9D%B4%ED%84%B0-%EC%B6%94%EC%B6%9C


pyautogui.alert('check')

file = "C:/study/CS codes.xlsx"
#import an existing Excel file
wb = openpyxl.load_workbook(file)

ws = wb["Sheet1"]
# ws = wb["nicknames with CS codes"]
#the below function counts even empty cell, it leads unexpect result
# col_max = ws.max_column
# max_row = ws.max_row

# col_max = max(len(row) for row in ws.values)
# max_row = len(list(ws.values))


#https://stackoverflow.com/questions/46569496/openpyxl-max-row-and-max-column-wrongly-reports-a-larger-figure
# print(col_max)
# print(max_row)
#ws.cell(row=1, column = col_max, value='happy')

# ws.cell(row=2, column =1, value="2030-01-01")
# ws.append(["2030-01-02", "pen", 400, 5, "=C3*D3"])

#find the maximum value of rows
for max_row, row in enumerate(ws, 1):
    if all(c.value is None for c in row):
        break #If all cells in a row are empty, it breaks out of the loop.
print(max_row)


#find the maximum value of columns
max_col = 0
for col in ws.iter_cols(): #iterate over each column in the worksheet ws
    if any(cell.value is not None for cell in col): #check if at least one cell in that column has a non-empty value
        max_col += 1
print(max_col)
ws.cell(row=1, column = max_col+1, value='happy')

# save file
wb.save(file)