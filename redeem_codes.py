from selenium import webdriver
import time
from selenium.webdriver.common.by import By
import pandas as pd  #pip install pandas
import pyautogui as pyautogui #pip install pyautogui
import openpyxl
import sys
import os


#파일 이름 본인이 정하게 하기? 경로는 어디로하지..
file = "C:/study/CS codes.xlsx"

previous_check =pyautogui.confirm("파일경로에 전용양식(CS codes.xlsx)에 모든 정보를 입력하셨나요?", buttons=['yes', 'no'])


if previous_check == "yes":
    pyautogui.alert("쿠폰 등록을 시작합니다.")
else:
    check_existence = os.path.exists(file)
    if check_existence == True:
        check_overwriting =pyautogui.confirm("이미 파일이 존재합니다. 덮어씌우시겠습니까?", buttons=['yes', 'no'])
        if check_overwriting == "yes":
            wb = openpyxl.Workbook()
            #select currently active sheet
            ws = wb.active

            #change name of sheet
            ws.title = "nicknames with CS codes"

            #write the head
            ws.append(["nickname", "cs code", "이 셀의 내용을 지우고 등록할 쿠폰코드를 써주세요!"])
            wb.save(file)

            pyautogui.alert("양식 파일 생성을 완료했습니다. 파일에 정보를 입력 후 다시 프로그램을 실행해주세요.")

            os.startfile(file) #open the Excel file
            
        else:
             pyautogui.alert("파일생성을 중단합니다.")
        sys.exit()  # forcefully exiting the program



driver = webdriver.Chrome()

driver.get("https://coupon.withhive.com/720")

#time.sleep(second): Suspend execution of the calling thread for the given number of seconds.
#time.sleep(10) #Wait 10 seconds for the page to completely load

##record the result in the Excel file
#import the Excel file
wb = openpyxl.load_workbook(file)
ws = wb["nicknames with CS codes"]

#find the maximum value of rows
for max_row, row in enumerate(ws, 1):
    if all(c.value is None for c in row):
        break #If all cells in a row are empty, it breaks out of the loop.

#find the maximum value of columns
max_col = 0
for col in ws.iter_cols(): #iterate over each column in the worksheet ws
    if any(cell.value is not None for cell in col): #check if at least one cell in that column has a non-empty value
        max_col += 1



##
#get CS codes in the excel file
df = pd.read_excel(file, sheet_name="nicknames with CS codes")

#select the cs code column(its index is 1)
cs_codes = df.iloc[:,[1]]

#.dropna(): remove NaN values
#.values.tolist(): convert as a list
cs_codes = cs_codes.dropna().values.tolist() 

#convert the data type to int
#[105548.0] -> 105548
cs_codes = [int(float(str(x).split('.')[0])) for sublist in cs_codes for x in sublist]


#coupon_code = df.iat[0, max_col]이라고 하면 에러남
coupon_code = df.columns[max_col-1]

#make list to contain index of cs codes that successed 
complete = []

#make dictionary to contain index of cs codes that failed and that reason
fail = {} 

for cs_code in cs_codes:
        
    #select a server
    driver.find_element(By.CSS_SELECTOR, "#HIVEcoupon > div.content > div > div.input_box.server_list_area.show > div.select_wrap > button").click()
    driver.find_element(By.CSS_SELECTOR, "#HIVEcoupon > div.content > div > div.input_box.server_list_area.show > div.select_wrap > ul > li > button").click()

    #input CS code and coupon code 
    driver.find_element(By.CSS_SELECTOR, "#cs_code").send_keys(cs_code)
    driver.find_element(By.CSS_SELECTOR, "#coupon_code").send_keys(coupon_code)
    driver.find_element(By.CSS_SELECTOR, "#HIVEcoupon > div.content > div > button").click()

    time.sleep(1)

    #select a server again
    driver.find_element(By.CSS_SELECTOR, "body > div.pop_wrap.server.coupon_server_lyr > div > ul > li > label > span").click()
    driver.find_element(By.CSS_SELECTOR, "body > div.pop_wrap.server.coupon_server_lyr > div > div.btns > button.btn_confirm").click()

    #read the message of alert
    result = driver.find_element(By.CSS_SELECTOR, "#layer_msg").text

    if result.find("!") == -1: #when the message contain a letter '!', it means redeeming was completed.
        #if redeeming failed since the coupon was already used, append the cs_codes and that reason in 'fail'
        fail[cs_codes.index(cs_code)] = result
    else: #if the code was redeemed successfully, append the cs_codes in 'complete'
        complete.append(cs_codes.index(cs_code))

    #click the close button
    driver.find_element(By.CSS_SELECTOR, "#layer_close_btn").click()
    
    #refre a webpage
    driver.refresh()

driver.quit()




#record the result on the excel file.
for index in complete:
    ws.cell(row=index+2, column =max_col, value="등록 성공")

for index in fail:
    ws.cell(row=index+2, column =max_col, value=f"등록 실패({fail[index]})")

##엑셀열려있어서 오류나는 것 대비하기.
try:
    # save the file
    wb.save(file)
except PermissionError:
    pyautogui.alert("엑셀 파일을 닫아주세요")



##alert the result to the user
message = f"쿠폰 등록 성공: {len(complete)}명\n쿠폰 등록 실패: {len(fail)}명\n엑셀 참조"
pyautogui.alert(message, title='미겜천 쿠폰등록하기')